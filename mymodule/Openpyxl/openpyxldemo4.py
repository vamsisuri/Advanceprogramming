import openpyxl as xl
wb=xl.load_workbook("C://mydataset//studentdata.xlsx")
sheet=wb['Sheet4']
dict = {}
list = []
for i in range(1,4):
    list1 = []
for j in range(1,4):
    list.append(sheet.cell(j, i).value)
    dict[(sheet.cell(1,i).value)] = list
print(dict)
