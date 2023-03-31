import openpyxl
wb=openpyxl.load_workbook("c://mydataset//studentdata.xlsx")
sheet1=wb["Sheet1"]
print(sheet1.max_row)
print(sheet1.max_column)

excel_data=[]
for i in range(1,5):
    excel_data.append(sheet1.cell(i,1).value)
    print(excel_data)
