import openpyxl as xl
from openpyxl import Workbook
import pandas as pd
def Check_data(df,df1):
    wb=xl.load_workbook("C://mydataset//master_data.xlsx")
    wb=xl.load_workbook("C://mydataset//daily_data.xlsx")
    sheet = wb['data']
    sheet1 = wb['data']
    ws=wb.active
    for rows in range(2,22):
        for coloumn in range(2,3):
            value_daily=int(sheet.cell(rows, coloumn).value)
            for i in range(6,7):
                for j in range(2,22):
                    value_master=int(sheet1.cell(i,j).value)
                    total= value_daily+value_master
                    print(total)
Check_data(df=pd.read_excel("C://mydataset/daily_data.xlsx"),df1=pd.read_excel("c://mydataset//master_data.xlsx")1)





